from src.infra.Mongo.infra import MongoInfrastructure
from decouple import config
import openpyxl


class ServicesOpenPyXl:

    database = MongoInfrastructure.get_db_and_collection()
    list_projects = list(database.find())
    plan_xls = config('book')
    book = openpyxl.load_workbook(f'{plan_xls}', data_only=True)
    page = book[''] # Adicionar o nome da planilha desejada que se encontra dentro do arquivo .xls
    projects_dict = dict()
    maps = {}

    '''A lib "openpyxl" não foi testado com o uso do decouple, pois foi adicionado em refatoramento após o uso do 
     sistema, caso ocorra algum erro, apenas retirar o uso dela com a lib'''

    @classmethod
    def insert_data_plan(cls):
        row = ''  #Colocar um INT referente a linha em que a planilha  comeca a apresentar os dados.
        while add := cls.page.cell(row=row, column=1).value:
            cls.maps[add.lower()] = {'row': row, 'column': row - '' if row < '' else row}
            row += 1 # No valor do While foi colocado coordenadas referente a planilha que estava sendo usada, para mapear dados.

        for project in cls.list_projects:
            system_name = project.get("nome_do_sistema")
            cls.projects_dict[system_name] = project.get("dependencias")

        for project_db, dependencies in cls.projects_dict.items():
            position_project = cls.maps.get(project_db.lower())

            if position_project is None:
                print('does not exist', project_db)
                continue

            for project_plan in dependencies:
                position_plan = cls.maps.get(project_plan.lower())

                if position_plan is None:
                    print('does not exist', project_plan)
                    continue

                cls.page.cell(row=position_project['row'], column=position_plan['column']).value = 'x'
                cls.page.cell(row=position_plan['row'], column=position_project['column']).value = 'x'
                print('Success')
                pass

        cls.book.save('') # Dar nome para a Nova planilha com alterações
